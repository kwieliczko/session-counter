##################################################################################################									
#  Session sounter
#  
#  Created 27-11-2019 by Karol Wieliczko    karol.wieliczko () protonmail.com
##################################################################################################


import logging, sys
from cfg import config
from modules import db_oracle
from modules import functions
from builtins import staticmethod
import argparse
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import datetime
#from scipy.constants.constants import year
import argparse


# Runtime class inheriting from GeneralFunctions
class main(functions.GeneralFunctions):
    
    @staticmethod
    def run():
        
        db_name = str(sys.argv[1])
        
   
        if len(sys.argv)==1:
            print("Enter the name of the database to be tested. Example.:  python main.py sales")
            sys.exit(1)
        
            

        weekDays = ("Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday")
          
        # Excel
        workbook = Workbook()
        sheet = workbook.active
            
        # Object of class DbOracle
        db = db_oracle.DbOracle()

        # Connection to the main database of the program based on parameters from the 
		# 'cfg\config.py' file to download the connection string of the tested database
        conn_main = db.db_connect(config.Config().host_name, 
                                  config.Config().db_port, 
                                  config.Config().db_service_name, 
                                  config.Config().db_user, 
                                  config.Config().db_password)

        db_query_connection_strings  = 'SELECT * FROM CONNECTING_STRING where DB_SERVICE_NAME = \'' + db_name + '\''
        connection_string_tab  = db.get_query_result(conn_main, db_query_connection_strings, 6) 

        # connection to the examined bases
        conn_db = db.db_connect(connection_string_tab[1], 
                                connection_string_tab[2], 
                                connection_string_tab[3], 
                                connection_string_tab[4], 
                                connection_string_tab[5])
        
        
        # From the dba_hist_snapshot View, I create a list containing the last 7 days 
		# of call history along with dates, snapshot times and the number of calls in a given hour
        #
        #  w stylu:
        #
        #        Date       hour        the number of calls in a given hour
        #        2019-11-18 00:00:00    94
        #        2019-11-18 01:00:00    93
        #        2019-11-18 02:00:00    86
        #        2019-11-18 03:00:00    91
        
        count_connection_per_hour_tab = db.get_count_connection_per_hour_from_db(conn_db)


		# I am creating a temporary list containing only the number of calls in specific hours 
		# to check what is the max and what is the min
        # (it's about choosing the brightness of the color)
        
        connections_count_tab = []
        
        for row in count_connection_per_hour_tab:
            count_connections = row[1]
            connections_count_tab.append(count_connections)

        
        max_connenctions = max(connections_count_tab)
        min_connenctions = min(connections_count_tab)
        
           
        day_in_week = 7
        excel_row=1     
        hour  = 1
        data_old =''
        sheet_start = 'A'  
        number_of_hours_in_week = 168
        
        for row in  count_connection_per_hour_tab:
            
            if number_of_hours_in_week <= 0:
               break
                
                
            data = row[0]
            
            year = int(data[0:4])
            mount = int(data[5:7])
            day = int(data[8:10])
            
            thisDay    = datetime.date(year,mount,day)
            thisWeekDay = thisDay.weekday()
            thisWeekDayAsString = weekDays[thisWeekDay]
                        
            # If it is another day we change the column and give the first row
            if data_old != data[0:10]: 
                
                # Header 
                sheet[''+ chr(ord(''+ sheet_start +'' ) + 2 ) +'1'] = thisWeekDayAsString
                sheet[''+ chr(ord(''+ sheet_start +'')  + 2 ) +'1'].fill = PatternFill(fgColor='999999', fill_type = "solid")     # Koloruje polenaglowka

                # Increase letter and column by 1
                sheet_start = chr(ord(sheet_start) + 1)
                excel_row = 2
                                
                
            
            
            # I insert an hour into Excel on the left, but only on the first day (in the first 24 hours)
            if hour <= 24:
                sheet[''+ chr(ord(''+ sheet_start +'') + 0) +''+str(excel_row)] = data[10:13]     # Hour
            
            sheet[''+ chr(ord(''+ sheet_start +'') + 1) +''+str(excel_row)] = int(row[1])     # Number of calls
            
            foreground_cell_color = functions.GeneralFunctions.choose_color(row[1], min_connenctions, max_connenctions)     # Dobieram kolor wypelnienia
            sheet[''+ chr(ord(''+ sheet_start +'') + 1) +''+str(excel_row)].fill = PatternFill(fgColor=foreground_cell_color, fill_type = "solid")     # Koloruje pole w Excelu
            
            excel_row = excel_row + 1
            hour = hour + 1
            
            data_old = data[0:10]
            
            number_of_hours_in_week = number_of_hours_in_week - 1
         
    

        # Save to Excel
        try:
            workbook.save(filename=db_name+'.xlsx')
            print("Excel generated")
            
        except:
            functions.GeneralFunctions.print_error("Error generating Excel file")
            return '0'
            pass
        
       
        
        # Close connection to the base
        conn_main.close()


# Starting the program
main().run()    



